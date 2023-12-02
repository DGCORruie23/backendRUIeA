from django.shortcuts import render
from django.db.models import Max, Count
from django.views.decorators.csrf import csrf_exempt
from rest_framework.parsers import JSONParser
from django.http import JsonResponse, HttpResponseRedirect, HttpResponse
from django.contrib.auth.hashers import make_password, check_password
from .serializers import UserGetSerializer, UserGetSerializerC, PaisesGetSerializer, EstadoFuerzaGetSerializer, FrasesGetSerializer, ListRescatePuntoSerializer, RescatePuntoSerializer
from .serializers import MunicipiosGetSerializer, PuntosInterGetSerializer, ConteoRapidoSerializer
from .models import Usuario, Paises, EstadoFuerza, Frases, Municipios, PuntosInternacion, RescatePunto, ConteoRapidoPunto
from .forms import CargarArchivoForm, ExcelForm, ExcelFormOr, ExcelFormOrs
import openpyxl as opxl
from openpyxl.writer.excel import save_virtual_workbook
import datetime

# Create your views here.

@csrf_exempt
def login_user(request):

    if request.method == 'POST':
        datos_recibidos = JSONParser().parse(request)
        datos_serializados = UserGetSerializer(data=datos_recibidos)
        # print(datos_serializados)
        if datos_serializados.is_valid():
            datos = datos_serializados.data
            usuario_info = datos.get('nickname')
            pass_info = datos.get('password')

            # print(f" nickname: {usuario_info}, pass: {pass_info}")

            if(Usuario.objects.filter(nickname = usuario_info).exists()):

                datos_usuario = Usuario.objects.get(nickname = usuario_info)

                if(check_password(pass_info, datos_usuario.password)):
                    nuevo_serializer = {'nickname': datos_usuario.nickname,
                                        'nombre': datos_usuario.nombre,
                                        'apellido': datos_usuario.apellido,
                                        'password': 'ok',
                                        'estado': datos_usuario.estado,
                                        'tipo': datos_usuario.tipo}
                    
                    # print(f" ${datos_usuario.nickname} ${datos_usuario.nombre} ${datos_usuario.apellido} ${datos_usuario.estado}")
                    serializer_datos = UserGetSerializerC(nuevo_serializer)
                    return JsonResponse(serializer_datos.data, status = 200)

                else:
                    nuevo_serializer = {'nickname': 'error',
                                        'nombre': '',
                                        'apellido': '',
                                         'password': 'error',
                                         'estado': '',
                                         'tipo': ''}
                    serializer_datos = UserGetSerializerC(nuevo_serializer)
                    return JsonResponse(serializer_datos.data, status = 200)

            else:

                nuevo_serializer = {'nickname': 'error',
                                        'nombre': '',
                                        'apellido': '',
                                         'password': 'error',
                                         'estado': '',
                                         'tipo': ''}
                serializer_datos = UserGetSerializerC(nuevo_serializer)
                return JsonResponse(serializer_datos.data, status = 200)
        else:
            nuevo_serializer = {'nickname': 'error',
                                        'nombre': '',
                                        'apellido': '',
                                         'password': 'error',
                                         'estado': '',
                                         'tipo': ''}
            serializer_datos = UserGetSerializerC(nuevo_serializer)
            return JsonResponse(serializer_datos.data, status = 200)

@csrf_exempt
def cargarPais(request):
    if(request.method == "POST"):
        form = CargarArchivoForm(request.POST, request.FILES)
        if(form.is_valid()):
            # Nota: "archivo" este campo se llama como se llama en el form  
            excel_file = request.FILES["archivo"]
            nombreA = str(excel_file.name)
            extensionA = (nombreA.split(".")[-1]).lower()
            # print((nombreA.split(".")[-1]).lower())
            if( extensionA == "xlsx" 
               or extensionA == ".xlsm" 
               or extensionA == ".xlsb" 
               or extensionA == ".xltx" 
               or extensionA == ".xltm" 
               or extensionA == ".xls"):
                dataWB = opxl.load_workbook(excel_file, data_only=True)

                data = dataWB.worksheets[0]

                # print(data.cell(4,2).value)

                # Se leen los datos del excel
                paises = []
                i = 1
                while not(i == 0):
                    if(data.cell(i+3,2).value == None):
                        i = 0
                    else:
                        paises.append([data.cell(i+3,2).value, data.cell(i+3,3).value])
                        i+=1
                #-----------------------------------

                #Eliminar todos los registros de Paises
                Paises.objects.all().delete()
                #------------------------

                # Se agrega a la DB los paises y se crea un string de ellos
                paisesStr = "{"    
                for num, nac in enumerate(paises):
                    paisesStr += f'\n\t"{nac}",' 
                    Paises.objects.create(idPais = num + 1, nombre_pais = nac[0], iso3 = nac[1])
                    # print(nac)
                paisesStr += "\n}" 
                #----------------------------

                # pasar string de las nacionalidades 
                # print(paisesStr) 
                #------------------------
            return HttpResponseRedirect("/info/cargarPais")
    else:
        form = CargarArchivoForm()
    return render(request, "cargarExcel/cargarArchivoPaises.html", {"form" : form})

@csrf_exempt
def cargarPuntoI(request):
    if(request.method == "POST"):
        form = CargarArchivoForm(request.POST, request.FILES)
        if(form.is_valid()):
            # Nota: "archivo" este campo se llama como se llama en el form  
            excel_file = request.FILES["archivo"]
            nombreA = str(excel_file.name)
            extensionA = (nombreA.split(".")[-1]).lower()
            # print((nombreA.split(".")[-1]).lower())
            if( extensionA == "xlsx" 
               or extensionA == ".xlsm" 
               or extensionA == ".xlsb" 
               or extensionA == ".xltx" 
               or extensionA == ".xltm" 
               or extensionA == ".xls"):
                dataWB = opxl.load_workbook(excel_file, data_only=True)

                data = dataWB.worksheets[0]

                # print(data.cell(4,2).value)

                # Se leen los datos del excel
                puntoI = []
                i = 1
                while not(i == 0):
                    if(data.cell(i,1).value == None):
                        i = 0
                    else:
                        puntoI.append([data.cell(i,1).value, data.cell(i,2).value, data.cell(i,3).value])
                        i+=1
                #-----------------------------------

                #Eliminar todos los registros de Paises
                PuntosInternacion.objects.all().delete()
                #------------------------

                # Se agrega a la DB los paises y se crea un string de ellos
                puntoIStr = "{"    
                for num, punto in enumerate(puntoI):
                    puntoIStr += f'\n\t"{punto}",' 
                    PuntosInternacion.objects.create(idPuntoInter = num + 1, estadoPunto = punto[0], nombrePunto = punto[1], tipoPunto = punto[2], )
                    # print(nac)
                puntoIStr += "\n}" 
                #----------------------------

                # pasar string de las nacionalidades 
                # print(puntoIStr) 
                #------------------------
            return HttpResponseRedirect("/info/cargarPuntosI")
    else:
        form = CargarArchivoForm()
    return render(request, "cargarExcel/cargarArchivoPuntosInter.html", {"form" : form})


@csrf_exempt
def cargarEdoFuerza(request):
    if(request.method == "POST"):
        form = CargarArchivoForm(request.POST, request.FILES)
        if(form.is_valid()):
            # Nota: "archivo" este campo se llama como se llama en el form  
            excel_file = request.FILES["archivo"]
            nombreA = str(excel_file.name)
            extensionA = (nombreA.split(".")[-1]).lower()
            # print((nombreA.split(".")[-1]).lower())
            if( extensionA == "xlsx" 
               or extensionA == ".xlsm" 
               or extensionA == ".xlsb" 
               or extensionA == ".xltx" 
               or extensionA == ".xltm" 
               or extensionA == ".xls"):
                dataWB = opxl.load_workbook(excel_file, data_only=True)

                data = dataWB.worksheets[0]

                # print(data.cell(4,2).value)

                # Se leen los datos del excel
                edoFuerza = []
                auxL = []
            
                i = 1
                while not(i == 0):

                    auxL.clear()

                    if(data.cell(i+2,1).value == None):
                        i = 0
                        # print(edoFuerza)
                    else:
                        for x in range(1, 14, 1):
                            
                            ax = data.cell(i+2, x).value
                            
                            if(x == 13):
                                if(ax == "CECO RIO BRAVO"):
                                    auxL.append(1)
                                elif(ax == "CECO SUCHIATE"):
                                    auxL.append(2)
                                else:
                                    auxL.append(3)
                            else:
                                if(ax == None or ax == "N/A"):
                                    auxL.append("0")
                                else:
                                    auxL.append(ax)

                        edoFuerza.append(auxL[:])
                        i+=1
                #-----------------------------------

                

                #Verificar si la lista esta vacia para no guardar información
                if(bool(edoFuerza)):

                    # print(edoFuerza)

                    #Eliminar todos los registros de Paises
                    EstadoFuerza.objects.all().delete()
                    #------------------------

                    # Se agrega a la DB los paises y se crea un string de ellos
                    edoFuerzaStr = "{"    
                    for num, fuerzaP in enumerate(edoFuerza):
                        
                        edoFuerzaStr += "\n\t"
                        
                        edoFuerzaStr += f"{fuerzaP}"
                        
                        textoC, lat, lon = convertirU(fuerzaP[5])

                        EstadoFuerza.objects.create(
                            idEdoFuerza = num + 1 ,
                            oficinaR = str(fuerzaP[0]),
                            numPunto = int(str(fuerzaP[1])),
                            nomPuntoRevision = str(fuerzaP[2]),
                            tipoP = str(fuerzaP[3]),
                            ubicacion = str(fuerzaP[4]),
                            coordenadasTexto = textoC,
                            latitud = float(lat),
                            longitud = float(lon),
                            personalINM = int(str(fuerzaP[6])),
                            personalSEDENA = int(str(fuerzaP[7])),
                            personalMarina = int(str(fuerzaP[8])),
                            personalGuardiaN = int(str(fuerzaP[9])),
                            personalOTROS = int(str(fuerzaP[10])),
                            vehiculos = int(str(fuerzaP[11])),
                            seccion = int(str(fuerzaP[12])),
                            )
                        # print(nac)
                    edoFuerzaStr += "\n}" 
                    #----------------------------

                    # pasar string de las nacionalidades 
                    # print(edoFuerzaStr) 
                    #------------------------
            return HttpResponseRedirect("/info/cargarFuerza")
    else:
        form = CargarArchivoForm()
    return render(request, "cargarExcel/cargarArchivoFuerza.html", {"form" : form})

@csrf_exempt
def cargarMunicipios(request):
    if(request.method == "POST"):
        form = CargarArchivoForm(request.POST, request.FILES)
        if(form.is_valid()):
            # Nota: "archivo" este campo se llama como se llama en el form  
            excel_file = request.FILES["archivo"]
            nombreA = str(excel_file.name)
            extensionA = (nombreA.split(".")[-1]).lower()
            # print((nombreA.split(".")[-1]).lower())
            if( extensionA == "xlsx" 
               or extensionA == ".xlsm" 
               or extensionA == ".xlsb" 
               or extensionA == ".xltx" 
               or extensionA == ".xltm" 
               or extensionA == ".xls"):
                dataWB = opxl.load_workbook(excel_file, data_only=True)

                data = dataWB.worksheets[0]

                # print(data.cell(4,2).value)

                # Se leen los datos del excel
                municipio = []
                auxL = []
            
                i = 1
                while not(i == 0):

                    auxL.clear()

                    if(data.cell( i+4, 4).value == None):
                        i = 0
                        # print(edoFuerza)
                    else:
                        edo = data.cell(i + 3, 4).value
                        edoAbrev = data.cell(i + 4, 4).value
                        nomMun = data.cell(i + 4, 6).value

                        abrevMayus = ((edoAbrev.upper()).replace(".","")).replace(" ","")

                        if(abrevMayus == "AGS"):
                            edo = "AGUASCALIENTES"
                        elif(abrevMayus == "BC"):
                            edo = "BAJA CALIFORNIA"
                        elif(abrevMayus == "BCS"):
                            edo = "BAJA CALIFORNIA SUR"
                        elif(abrevMayus == "CAMP"):
                            edo = "CAMPECHE"
                        elif(abrevMayus == "CDMX"):
                            edo = "CDMX"
                        elif(abrevMayus == "CHIH"):
                            edo = "CHIHUAHUA"
                        elif(abrevMayus == "CHIS"):
                            edo = "CHIAPAS"
                        elif(abrevMayus == "COAH"):
                            edo = "COAHUILA"
                        elif(abrevMayus == "COL"):
                            edo = "COLIMA"
                        elif(abrevMayus == "DGO"):
                            edo = "DURANGO"
                        elif(abrevMayus == "GRO"):
                            edo = "GUERRERO"
                        elif(abrevMayus == "GTO"):
                            edo = "GUANAJUATO"
                        elif(abrevMayus == "HGO"):
                            edo = "HIDALGO"
                        elif(abrevMayus == "JAL"):
                            edo = "JALISCO"
                        elif(abrevMayus == "MEX"):
                            edo = "EDOMEX"
                        elif(abrevMayus == "MICH"):
                            edo = "MICHOACÁN"
                        elif(abrevMayus == "MOR"):
                            edo = "MORELOS"
                        elif(abrevMayus == "NAY"):
                            edo = "NAYARIT"
                        elif(abrevMayus == "NL"):
                            edo = "NUEVO LEÓN"
                        elif(abrevMayus == "OAX"):
                            edo = "OAXACA"
                        elif(abrevMayus == "PUE"):
                            edo = "PUEBLA"
                        elif(abrevMayus == "QROO"):
                            edo = "QUINTANA ROO"
                        elif(abrevMayus == "QRO"):
                            edo = "QUERÉTARO"
                        elif(abrevMayus == "SIN"):
                            edo = "SINALOA"
                        elif(abrevMayus == "SLP"):
                            edo = "SAN LUIS POTOSÍ"
                        elif(abrevMayus == "SON"):
                            edo = "SONORA"
                        elif(abrevMayus == "TAB"):
                            edo = "TABASCO"
                        elif(abrevMayus == "TAMPS"):
                            edo = "TAMAULIPAS"
                        elif(abrevMayus == "TLAX"):
                            edo = "TLAXCALA"
                        elif(abrevMayus == "VER"):
                            edo = "VERACRUZ"
                        elif(abrevMayus == "YUC"):
                            edo = "YUCATÁN"
                        elif(abrevMayus == "ZAC"):
                            edo = "ZACATECAS"
                        else:
                            edo = ""
                        
                        municipio.append([edo, abrevMayus, nomMun])

                        i+=1
                #-----------------------------------

                

                #Verificar si la lista esta vacia para no guardar información
                if(bool(municipio)):

                    # print(edoFuerza)

                    #Eliminar todos los registros de Paises
                    Municipios.objects.all().delete()
                    #------------------------

                    # Se agrega a la DB los paises y se crea un string de ellos
                    municipioStr = "{"    
                    for num, fuerzaP in enumerate(municipio):
                        
                        municipioStr += "\n\t"
                        
                        municipioStr += f"{fuerzaP}"

                        Municipios.objects.create(
                            idMunicipio = num + 1 ,
                            estado = str(fuerzaP[0]),
                            estadoAbr = str(fuerzaP[1]),
                            nomMunicipio = str(fuerzaP[2]),
                            )
                        # print(nac)
                    municipioStr += "\n}" 

                    # print("\n\nmunicipios")
                    # datosMun = Municipios.objects.values('estadoAbr').order_by('estadoAbr').distinct() 
                    # for x in datosMun:
                    #     print(x)

                    # print("\n\nOrs")
                    # datosOr = EstadoFuerza.objects.values('oficinaR').order_by('oficinaR').distinct() 
                    # for x in datosOr:
                    #     print(x)
                    #----------------------------

                    # pasar string de las nacionalidades 
                    # print(edoFuerzaStr) 
                    #------------------------
            return HttpResponseRedirect("/info/cargarMunicipios")
    else:
        form = CargarArchivoForm()
    return render(request, "cargarExcel/cargarArchivoMunicipios.html", {"form" : form})


@csrf_exempt
def insert_rescates(request):

    if request.method == 'POST':
        datos_recibidos = JSONParser().parse(request)
        datos_serializados = RescatePuntoSerializer(data=datos_recibidos, many=True)
        if datos_serializados.is_valid():
            datos_serializados.save()

            # print(datos)
            dataR = {
                'guardado' : "ok"
            }
            return JsonResponse(dataR, status = 200)
            
        else:
            print(datos_serializados.data)
            datos_serializados.is_valid
            print(datos_serializados.errors)
            dataR = {
                'guardado' : "error"
            }
            return JsonResponse( dataR, status = 200)

@csrf_exempt
def insert_conteo(request):

    if request.method == 'POST':
        datos_recibidos = JSONParser().parse(request)
        datos_serializados = ConteoRapidoSerializer(data=datos_recibidos, many=True)
        if datos_serializados.is_valid():
            datos_serializados.save()
            # print(datos)
            dataR = {
                'guardado' : "ok"
            }
            return JsonResponse(dataR, status = 200)
            
        else:
            # print(datos_serializados.data)
            # datos_serializados.is_valid
            # print(datos_serializados.errors)
            dataR = {
                'guardado' : "error"
            }
            return JsonResponse( dataR, status = 200)

@csrf_exempt
def infoPaises(request):
    if request.method == 'GET':
        snippets = Paises.objects.all()
        serializer = PaisesGetSerializer(snippets, many=True)
        return JsonResponse(serializer.data, safe=False)

@csrf_exempt
def infoMunicipios(request):
    if request.method == 'GET':
        snippets = Municipios.objects.all()
        serializer = MunicipiosGetSerializer(snippets, many=True)
        return JsonResponse(serializer.data, safe=False)
    
@csrf_exempt
def infoPuntosInterna(request):
    if request.method == 'GET':
        snippets = PuntosInternacion.objects.all()
        serializer = PuntosInterGetSerializer(snippets, many=True)
        return JsonResponse(serializer.data, safe=False)
    
@csrf_exempt
def infoEstadoFuerza(request):
    if request.method == 'GET':
        snippets = EstadoFuerza.objects.all()
        serializer = EstadoFuerzaGetSerializer(snippets, many=True)
        return JsonResponse(serializer.data, safe=False)
    
@csrf_exempt
def infoFrases(request):
    if request.method == 'GET':
        snippets = Frases.objects.all()
        serializer = FrasesGetSerializer(snippets, many=True)
        return JsonResponse(serializer.data, safe=False)


def convertirU(coordenadas):
    strCoor = str(coordenadas)
    lat = float("0.0")
    lon = float("0.0")
    return strCoor, lat, lon


@csrf_exempt
def generarExcelNombres(request):
    if(request.method == "POST"):
        form = ExcelForm(request.POST)
        if(form.is_valid()):
            # print(request.POST)
            # Nota: "archivo" este campo se llama como se llama en el form  
            dia = request.POST["fechaDescarga_day"]
            mes = request.POST["fechaDescarga_month"]
            year = request.POST["fechaDescarga_year"]

            fechaR = datetime.datetime.strptime(f"{dia}/{mes}/{year}", "%d/%m/%Y").strftime('%d-%m-%y')

            valores = RescatePunto.objects.filter(fecha= fechaR)
            
            workbook = opxl.load_workbook('tmp/dup.xlsm', read_only=False, keep_vba=True)
            worksheet = workbook.active
            
            # worksheet.append(['Oficina de Representación', 
            #                  'Fecha', 
            #                  'Hora',
            #                  'AGENTE',
            #                  'Aeropuerto', 
            #                  'Carretero', 
            #                  'Tipo de Vehículo', 
            #                  'Linea / Empresa', 
            #                  'No. Economico', 
            #                  'Placas', 
            #                  'Vehículo Asegurado', 
            #                  'Casa de Seguridad', 
            #                  'Central de Autobus', 
            #                  'Ferrocarril', 
            #                  'Empresa', 
            #                  'Hotel', 
            #                  'Nombre del Hotel', 
            #                  'Puestos a Disposición', 
            #                  'Juéz Calificador', 
            #                  'Reclusorio', 
            #                  'Policía Federal', 
            #                  'DIF', 
            #                  'Plicía Estatal', 
            #                  'Policía Municipal', 
            #                  'Guardia Nacional', 
            #                  'Fiscalia', 
            #                  'Otras Autoridades', 
            #                  'Voluntarios', 
            #                  'Otro', 
            #                  'Presuntos Delincuentes', 
            #                  'No. de Presuntos Delincuentes', 
            #                  'Municipio', 
            #                  'Punto Estratégico', 
            #                  'Nacionalidad', 
            #                  'ISO', 
            #                  'Nombre', 
            #                  'Apellidos',
            #                  'No de Documento',
            #                  'Parentesco',
            #                  'Fecha de Nacimiento',
            #                  'Sexo',
            #                  'Embarazada',
            #                  'No. de Familia',
            #                  'Edad',
            #                  ])
            for valor in valores:
                worksheet.append([valor.oficinaRepre, 
                                 valor.fecha,
                                 valor.hora,
                                 valor.nombreAgente.upper(),
                                 "1" if valor.aeropuerto else "",
                                 "1" if valor.carretero else "",
                                 valor.tipoVehic.upper(),
                                 valor.lineaAutobus.upper(),
                                 valor.numeroEcono.upper(),
                                 valor.placas.upper(),
                                 "1" if valor.vehiculoAseg else "",
                                 "1" if valor.casaSeguridad else "",
                                 "1" if valor.centralAutobus else "",
                                 "1" if valor.ferrocarril else "",
                                 valor.empresa,
                                 "1" if valor.hotel else "",
                                 valor.nombreHotel,
                                 "1" if valor.puestosADispo else "",
                                 "1" if valor.juezCalif else "",
                                 "1" if valor.reclusorio else "",
                                 "1" if valor.policiaFede else "",
                                 "1" if valor.dif else "",
                                 "1" if valor.policiaEsta else "",
                                 "1" if valor.policiaMuni else "",
                                 "1" if valor.guardiaNaci else "",
                                 "1" if valor.fiscalia else "",
                                 "1" if valor.otrasAuto else "",
                                 "1" if valor.voluntarios else "",
                                 "1" if valor.otro else "",
                                 "1" if valor.presuntosDelincuentes else "",
                                 valor.numPresuntosDelincuentes if valor.numPresuntosDelincuentes != 0 else "",
                                 valor.municipio.upper(),
                                 valor.puntoEstra.upper(),
                                 valor.nacionalidad.upper(),
                                 valor.iso3,
                                 valor.nombre.upper(),
                                 valor.apellidos.upper(),
                                 valor.noIdentidad,
                                 valor.parentesco.upper(),
                                 valor.fechaNacimiento,
                                 "Hombre" if valor.sexo else "Mujer",
                                 "1" if valor.embarazo else "",
                                 valor.numFamilia if valor.numFamilia != 0 else "",
                                 valor.edad,
                                 ])
            
            worksheet.append(['Rescates Totales: ', valores.count()])

            response = HttpResponse(content = save_virtual_workbook(workbook), content_type='application/vnd.ms-excel.sheet.macroEnabled.12')
            response['Content-Disposition'] = 'attachment; filename={fecha}.xlsm'.format(fecha = fechaR)

            return response
    else:
        form = ExcelForm()
    return render(request, "generarExcel/generarArchivoExcelNombres.html", {"form" : form})


@csrf_exempt
def generarExcelConteo(request):
    if(request.method == "POST"):
        form = ExcelForm(request.POST)
        print("entro")
        if(form.is_valid()):
            # print(request.POST)
            # Nota: "archivo" este campo se llama como se llama en el form  
            dia = request.POST["fechaDescarga_day"]
            mes = request.POST["fechaDescarga_month"]
            year = request.POST["fechaDescarga_year"]

            fechaR = datetime.datetime.strptime(f"{dia}/{mes}/{year}", "%d/%m/%Y").strftime('%d-%m-%y')

            valores = ConteoRapidoPunto.objects.filter(fecha= fechaR)
            
            workbook = opxl.load_workbook('tmp/dup.xlsm', read_only=False, keep_vba=True)
            worksheet = workbook.active

            # worksheet.append(['Eventos Totales: ', valores.count(), 'Adulto No Acompañado = ANA', '', '','Adulto Acompañado = AA'])

            # worksheet.append(['Oficina de Representación', 
            #                  'Fecha', 
            #                  'Hora',
            #                  'AGENTE',
            #                  'Aeropuerto', 
            #                  'Carretero', 
            #                  'Tipo de Vehículo', 
            #                  'Linea / Empresa', 
            #                  'No. Economico', 
            #                  'Placas', 
            #                  'Vehículo Asegurado', 
            #                  'Casa de Seguridad', 
            #                  'Central de Autobus', 
            #                  'Ferrocarril', 
            #                  'Empresa', 
            #                  'Hotel', 
            #                  'Nombre del Hotel', 
            #                  'Puestos a Disposición', 
            #                  'Juéz Calificador', 
            #                  'Reclusorio', 
            #                  'Policía Federal', 
            #                  'DIF', 
            #                  'Plicía Estatal', 
            #                  'Policía Municipal', 
            #                  'Guardia Nacional', 
            #                  'Fiscalia', 
            #                  'Otras Autoridades', 
            #                  'Voluntarios', 
            #                  'Otro', 
            #                  'Presuntos Delincuentes', 
            #                  'No. de Presuntos Delincuentes', 
            #                  'Municipio', 
            #                  'Punto Estratégico',

            #                  'Nacionalidad', 
            #                  'ISO', 
                             
            #                  'ANA Hombre(s)', 
            #                  'ANA Mujer(es)',
            #                  'ANA Mujer No Embarazada(s)',
                             
            #                  'Nucleos Familiares',
            #                  'AA Hombre(s)', 
            #                  'AA Mujer(es)',
            #                  'AA Mujer No Embarazada(s)',
            #                  'NNAA Hombre(s)', 
            #                  'NNAA Mujer(es)',
            #                  'NNAA Mujer No Embarazada(s)',
                             
            #                  'NNANA Hombre(s)', 
            #                  'NNANA Mujer(es)',
            #                  'NNANA Mujer No Embarazada(s)',
            #                  'Masivo'
            #                  ])
            for valor in valores:
                masivoD = (valor.AS_hombres + valor.AS_mujeresNoEmb +valor.AS_mujeresEmb +
                           valor.AA_hombres + valor.AA_mujeresNoEmb + valor.AA_mujeresEmb +
                           valor.NNA_A_hombres + valor.NNA_A_mujeresNoEmb + valor.NNA_A_mujeresEmb +
                           valor.NNA_S_hombres + valor.NNA_S_mujeresNoEmb + valor.NNA_S_mujeresEmb
                           )
                worksheet.append([valor.oficinaRepre, 
                                 valor.fecha,
                                 valor.hora,
                                 valor.nombreAgente.upper(),
                                 "1" if valor.aeropuerto else "",
                                 "1" if valor.carretero else "",
                                 valor.tipoVehic.upper(),
                                 valor.lineaAutobus.upper(),
                                 valor.numeroEcono.upper(),
                                 valor.placas.upper(),
                                 "1" if valor.vehiculoAseg else "",
                                 "1" if valor.casaSeguridad else "",
                                 "1" if valor.centralAutobus else "",
                                 "1" if valor.ferrocarril else "",
                                 valor.empresa,
                                 "1" if valor.hotel else "",
                                 valor.nombreHotel,
                                 "1" if valor.puestosADispo else "",
                                 "1" if valor.juezCalif else "",
                                 "1" if valor.reclusorio else "",
                                 "1" if valor.policiaFede else "",
                                 "1" if valor.dif else "",
                                 "1" if valor.policiaEsta else "",
                                 "1" if valor.policiaMuni else "",
                                 "1" if valor.guardiaNaci else "",
                                 "1" if valor.fiscalia else "",
                                 "1" if valor.otrasAuto else "",
                                 "1" if valor.voluntarios else "",
                                 "1" if valor.otro else "",
                                 "1" if valor.presuntosDelincuentes else "",
                                 valor.numPresuntosDelincuentes if valor.numPresuntosDelincuentes != 0 else "",
                                 valor.municipio.upper(),
                                 valor.puntoEstra.upper(),
                                 valor.nacionalidad.upper(),
                                 valor.iso3,

                                 valor.AS_hombres,
                                 valor.AS_mujeresNoEmb,
                                 valor.AS_mujeresEmb,

                                 valor.nucleosFamiliares,
                                 valor.AA_hombres,
                                 valor.AA_mujeresNoEmb,
                                 valor.AA_mujeresEmb,
                                 valor.NNA_A_hombres,
                                 valor.NNA_A_mujeresNoEmb,
                                 valor.NNA_A_mujeresEmb,

                                 valor.NNA_S_hombres,
                                 valor.NNA_S_mujeresNoEmb,
                                 valor.NNA_S_mujeresEmb,

                                 "1" if masivoD >= 40 else "",
                                 ])
            
            response = HttpResponse(content = save_virtual_workbook(workbook), content_type='application/vnd.ms-excel.sheet.macroEnabled.12')
            response['Content-Disposition'] = 'attachment; filename=rapido {fecha}.xlsm'.format(fecha = fechaR)
            return response
    else:
        form = ExcelForm()
    return render(request, "generarExcel/generarArchivoExcelConteo.html", {"form" : form})


def servirApps(request):
    return render(request, "descargas/descargar_Apps.html", {

    })

@csrf_exempt
def downloadAPK(request):
    if request.method == 'GET':
        appAndroid = open('tmp/ruie.apk', 'rb')

        response = HttpResponse(appAndroid, content_type="application/vnd.android.package-archive")
        response["Content-disposition"] = "attachment; filename=ruie.apk"
        return response
    
@csrf_exempt
def pagDuplicados(request):
    if request.method == 'GET':
        return render(request, "descargas/descargar_duplicados.html", {})
    
@csrf_exempt
def downloadDuplicados(request):
    if request.method == 'GET':

        workbook = opxl.load_workbook('tmp/dup.xlsm', read_only=False, keep_vba=True)
        worksheet = workbook.active

        fechaR = datetime.datetime.today().strftime('%d-%m-%y')

        register3 = RescatePunto.objects.values("iso3", "nombre", "apellidos", "fechaNacimiento").annotate(records=Count("*")).filter(records__gt=1)
        duplicados = []
        for reg in register3:
            for registro in RescatePunto.objects.filter(iso3=reg['iso3'], nombre=reg['nombre'], apellidos = reg['apellidos'], fechaNacimiento=reg['fechaNacimiento']):
                duplicados.append(registro)
                
        for valor in duplicados:
            worksheet.append([valor.oficinaRepre, 
                              valor.fecha,
                              valor.hora,
                              valor.nombreAgente.upper(),
                              "1" if valor.aeropuerto else "",
                              "1" if valor.carretero else "",
                              valor.tipoVehic.upper(),
                              valor.lineaAutobus.upper(),
                              valor.numeroEcono.upper(),
                              valor.placas.upper(),
                              "1" if valor.vehiculoAseg else "",
                              "1" if valor.casaSeguridad else "",
                              "1" if valor.centralAutobus else "",
                              "1" if valor.ferrocarril else "",
                              valor.empresa,
                              "1" if valor.hotel else "",
                              valor.nombreHotel,
                              "1" if valor.puestosADispo else "",
                              "1" if valor.juezCalif else "",
                              "1" if valor.reclusorio else "",
                              "1" if valor.policiaFede else "",
                              "1" if valor.dif else "",
                              "1" if valor.policiaEsta else "",
                              "1" if valor.policiaMuni else "",
                              "1" if valor.guardiaNaci else "",
                              "1" if valor.fiscalia else "",
                              "1" if valor.otrasAuto else "",
                              "1" if valor.voluntarios else "",
                              "1" if valor.otro else "",
                              "1" if valor.presuntosDelincuentes else "",
                              valor.numPresuntosDelincuentes if valor.numPresuntosDelincuentes != 0 else "",
                              valor.municipio.upper(),valor.puntoEstra.upper(),
                              valor.nacionalidad.upper(),
                              valor.iso3,
                              valor.nombre.upper(),
                              valor.apellidos.upper(),
                              valor.noIdentidad,
                              valor.parentesco.upper(),
                              valor.fechaNacimiento,
                              "Hombre" if valor.sexo else "Mujer",
                              "1" if valor.embarazo else "",
                              valor.numFamilia if valor.numFamilia != 0 else "",
                              valor.edad,
                              ])

        response = HttpResponse(content = save_virtual_workbook(workbook), content_type='application/vnd.ms-excel.sheet.macroEnabled.12')
        response['Content-Disposition'] = 'attachment; filename={fecha}.xlsm'.format(fecha = fechaR)

        return response


@csrf_exempt
def generarExcelTab(request):
    if(request.method == "POST"):
        form = ExcelFormOr(request.POST)
        if(form.is_valid()):
            dia = request.POST["fechaDescarga_day"]
            mes = request.POST["fechaDescarga_month"]
            year = request.POST["fechaDescarga_year"]

            fechaR = datetime.datetime.strptime(f"{dia}/{mes}/{year}", "%d/%m/%Y").strftime('%d-%m-%y')

            valores = valores = RescatePunto.objects.filter(fecha= fechaR).filter( oficinaRepre="TABASCO")
            
            workbook = opxl.load_workbook('tmp/dup.xlsm', read_only=False, keep_vba=True)
            worksheet = workbook.active
            
            for valor in valores:
                worksheet.append([valor.oficinaRepre, 
                                 valor.fecha,
                                 valor.hora,
                                 valor.nombreAgente.upper(),
                                 "1" if valor.aeropuerto else "",
                                 "1" if valor.carretero else "",
                                 valor.tipoVehic.upper(),
                                 valor.lineaAutobus.upper(),
                                 valor.numeroEcono.upper(),
                                 valor.placas.upper(),
                                 "1" if valor.vehiculoAseg else "",
                                 "1" if valor.casaSeguridad else "",
                                 "1" if valor.centralAutobus else "",
                                 "1" if valor.ferrocarril else "",
                                 valor.empresa,
                                 "1" if valor.hotel else "",
                                 valor.nombreHotel,
                                 "1" if valor.puestosADispo else "",
                                 "1" if valor.juezCalif else "",
                                 "1" if valor.reclusorio else "",
                                 "1" if valor.policiaFede else "",
                                 "1" if valor.dif else "",
                                 "1" if valor.policiaEsta else "",
                                 "1" if valor.policiaMuni else "",
                                 "1" if valor.guardiaNaci else "",
                                 "1" if valor.fiscalia else "",
                                 "1" if valor.otrasAuto else "",
                                 "1" if valor.voluntarios else "",
                                 "1" if valor.otro else "",
                                 "1" if valor.presuntosDelincuentes else "",
                                 valor.numPresuntosDelincuentes if valor.numPresuntosDelincuentes != 0 else "",
                                 valor.municipio.upper(),
                                 valor.puntoEstra.upper(),
                                 valor.nacionalidad.upper(),
                                 valor.iso3,
                                 valor.nombre.upper(),
                                 valor.apellidos.upper(),
                                 valor.noIdentidad,
                                 valor.parentesco.upper(),
                                 valor.fechaNacimiento,
                                 "Hombre" if valor.sexo else "Mujer",
                                 "1" if valor.embarazo else "",
                                 valor.numFamilia if valor.numFamilia != 0 else "",
                                 valor.edad,
                                 ])
            
            worksheet.append(['Rescates Totales: ', valores.count()])

            response = HttpResponse(content = save_virtual_workbook(workbook), content_type='application/vnd.ms-excel.sheet.macroEnabled.12')
            response['Content-Disposition'] = 'attachment; filename={fecha}.xlsm'.format(fecha = fechaR)

            return response
    else:
        form = ExcelFormOr()
    return render(request, "generarExcel/generarArchivoExcelOrs.html", {"form" : form})

@csrf_exempt
def generarExcelORs(request):
    if(request.method == "POST"):
        form = ExcelFormOrs(request.POST)
        # print(request.POST["fechaDescarga"])
        # print(request.POST["oficina"])
        # dia = request.POST["fechaDescarga_day"]
        # mes = request.POST["fechaDescarga_month"]
        # year = request.POST["fechaDescarga_year"]
        fechaR = request.POST["fechaDescarga"]

        oficinaR = request.POST["oficina"]

        # fechaR = datetime.datetime.strptime(f"{dia}/{mes}/{year}", "%d/%m/%Y").strftime('%d-%m-%y')
        # fechaR = datetime.datetime.strptime(f"{fechaFun}", "%d/%m/%Y").strftime('%d-%m-%y')

        valores = valores = RescatePunto.objects.filter(fecha= fechaR).filter( oficinaRepre=oficinaR)
        
        workbook = opxl.load_workbook('tmp/dup.xlsm', read_only=False, keep_vba=True)
        worksheet = workbook.active
            
        for valor in valores:
            worksheet.append([valor.oficinaRepre, 
                                valor.fecha,
                                valor.hora,
                                valor.nombreAgente.upper(),
                                "1" if valor.aeropuerto else "",
                                "1" if valor.carretero else "",
                                valor.tipoVehic.upper(),
                                valor.lineaAutobus.upper(),
                                valor.numeroEcono.upper(),
                                valor.placas.upper(),
                                "1" if valor.vehiculoAseg else "",
                                "1" if valor.casaSeguridad else "",
                                "1" if valor.centralAutobus else "",
                                "1" if valor.ferrocarril else "",
                                valor.empresa,
                                "1" if valor.hotel else "",
                                valor.nombreHotel,
                                "1" if valor.puestosADispo else "",
                                "1" if valor.juezCalif else "",
                                "1" if valor.reclusorio else "",
                                "1" if valor.policiaFede else "",
                                "1" if valor.dif else "",
                                "1" if valor.policiaEsta else "",
                                "1" if valor.policiaMuni else "",
                                "1" if valor.guardiaNaci else "",
                                "1" if valor.fiscalia else "",
                                "1" if valor.otrasAuto else "",
                                "1" if valor.voluntarios else "",
                                "1" if valor.otro else "",
                                "1" if valor.presuntosDelincuentes else "",
                                valor.numPresuntosDelincuentes if valor.numPresuntosDelincuentes != 0 else "",
                                valor.municipio.upper(),
                                valor.puntoEstra.upper(),
                                valor.nacionalidad.upper(),
                                valor.iso3,
                                valor.nombre.upper(),
                                valor.apellidos.upper(),
                                valor.noIdentidad,
                                valor.parentesco.upper(),
                                valor.fechaNacimiento,
                                "Hombre" if valor.sexo else "Mujer",
                                "1" if valor.embarazo else "",
                                valor.numFamilia if valor.numFamilia != 0 else "",
                                valor.edad,
                                ])
        
        worksheet.append(['Rescates Totales: ', valores.count()])
        worksheet.append(['Oficina: ', oficinaR])

        response = HttpResponse(content = save_virtual_workbook(workbook), content_type='application/vnd.ms-excel.sheet.macroEnabled.12')
        response['Content-Disposition'] = 'attachment; filename={fecha}.xlsm'.format(fecha = fechaR)

        return response
    
    else:
        return render(request, "base/error404.html")